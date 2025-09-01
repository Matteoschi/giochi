import os
from openpyxl import Workbook, load_workbook
import random
import json
from collections import Counter

territori = r"C:\Users\alessandrini\Documents\coding\games\Risiko\territori.json"
obiettivi = r"C:\Users\alessandrini\Documents\coding\games\Risiko\obiettivi.json"
continenti = r"C:\Users\alessandrini\Documents\coding\games\Risiko\continenti.json"
FOLDER = "Risiko"
os.makedirs(FOLDER, exist_ok=True)
NOME_FILE_EXCEL = "RisiKo.xlsx"
EXCEL_PATH = os.path.join(FOLDER, NOME_FILE_EXCEL)

# ------------------- CARICAMENTO FILE -------------------
def carica_obiettivi():
    try:
        if not os.path.exists(obiettivi):
            print(f"❌ File obiettivi non trovato: {obiettivi}")
            return []
        else:
            with open(obiettivi, 'r', encoding='utf-8') as file_obiettivi:
                return json.load(file_obiettivi)
    except json.JSONDecodeError:
        print("❌ Errore nel file obiettivi.json")
        return []


def carica_territori():
    try:
        if not os.path.exists(territori):
            print(f"❌ File territori non trovato: {territori}")
            return []
        else:
            with open(territori, 'r', encoding='utf-8') as file_territori:
                print("✅ File territori caricato con successo.")
                return json.load(file_territori)
    except json.JSONDecodeError:
        print("❌ Errore nel file territori.json")
        return []

def carica_continenti():
    try:
        if not os.path.exists(continenti):
            print(f"❌ File continenti non trovato: {continenti}")
            return []
        else:
            with open(continenti, 'r', encoding='utf-8') as file_continenti:
                print("✅ File continenti caricato con successo.")
                return json.load(file_continenti)
    except json.JSONDecodeError:
        print("❌ Errore nel file continenti.json")
        return []


# ------------------- CREAZIONE/VERIFICA FILE -------------------
def verifica_file(lista_giocatori, lista_colori, pedine_iniziali):
    try:
        if os.path.exists(EXCEL_PATH):
            eliminazione = input("attenzione la partita già esiste eliminarla ? (s/n)").lower().strip()
            if eliminazione == 's':
                print("📂 elimino file e creo nuova partita")
                os.remove(EXCEL_PATH)
            else:
                exit("❌ Operazione annullata, esco.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Giocatori"

        for giocatore, colore in zip(lista_giocatori, lista_colori):
            foglio = wb.create_sheet(title=giocatore)
            
            # Intestazione informazioni generali
            print(f"# Intestazione informazioni generali per {giocatore}")
            foglio["A1"] = "Giocatore"
            foglio["B1"] = giocatore
            foglio["A2"] = "Colore"
            foglio["B2"] = colore
            foglio["A3"] = "Pedine Iniziali"
            foglio["B3"] = pedine_iniziali
            foglio["A5"] = "OBIETTIVO SEGRETO"
            # Lasciamo B5 vuota, verrà riempita dalla funzione assegna_obiettivi

            print(f"# Intestazione territori per {giocatore}")
            foglio["A7"] = "Territori"
            foglio["A8"] = "Nome"
            foglio["B8"] = "Continente"
            foglio["C8"] = "Simbolo"
            foglio["D8"] = "numero Truppe"

        wb.save(EXCEL_PATH)
        print(f"✅ File '{EXCEL_PATH}' creato con i fogli dei giocatori.")
        return wb, ws

    except Exception as e:
        print(f"❌ Errore nel controllo dei file: {e}")
        return None, None

# ------------------- PEDINE INIZIALI -------------------

def pedine_distart(lista_giocatori):
    if len(lista_giocatori) == 3:
        return 35
    elif len(lista_giocatori) == 4:
        return 30
    elif len(lista_giocatori) == 5:
        return 25
    elif len(lista_giocatori) == 6:
        return 20

# ------------------- GIOCATORI -------------------
def aggiungi_giocatori():
    lista_giocatori = ["agata","matteo","asia","francesca","martina"]
    n_giocatori=len(lista_giocatori)
    print(f"\n🎮 Giocatori registrati: {lista_giocatori}")
    return lista_giocatori, n_giocatori

# ------------------- ASSEGNAZIONE COLORI -------------------
def assegna_colore(lista_giocatori):
    colori_disponibili = ["giallo", "rosso", "verde", "blu", "viola", "nero"]
    random.shuffle(colori_disponibili)
    print("\n🎨 Colori assegnati:")
    lista_colori = colori_disponibili[:len(lista_giocatori)]
    for giocatore, colore in zip(lista_giocatori, lista_colori):
        print(f" - {giocatore}: {colore}")
    return lista_colori

# ------------------- ASSEGNAZIONE TERRITORI -------------------
def assegna_territori(file_territori, lista_giocatori, wb):
    territori = file_territori.copy()
    random.shuffle(territori)

    # numero di territori per ciascun giocatore
    numero_carte = len(territori) // len(lista_giocatori)

    for giocatore in lista_giocatori:
        ws = wb[giocatore]

        for _ in range(numero_carte):
            territorio = territori.pop()
            ws.append([
                territorio["nome"].lower(),
                territorio["continente"].lower(),
                territorio["simbolo"].lower().strip()
            ])

    wb.save(EXCEL_PATH)
    print("✅ Territori assegnati ai giocatori, visulaizzabili file exel")
    return numero_carte

# ------------------- ASSEGNAZIONE OBIETTIVI -------------------

def assegna_obiettivi(lista_giocatori, wb, file_obiettivi):
    obiettivi= file_obiettivi.copy()
    random.shuffle(obiettivi)

    for giocatore in lista_giocatori:
        ws = wb[giocatore]
        obiettivo = obiettivi.pop()
        ws["B5"] = obiettivo["descrizione"]  # Obiettivo segreto in B5
        print(f"✅ Obiettivo per {giocatore} assegnato.")
    wb.save(EXCEL_PATH)
    print(" \n obiettivi assegnati ai giocatori, visulaizza file exel")

# ------------------- ASSEGNAZIONE TRUPPE -------------------

def inserire_truppe_iniziali(lista_giocatori,pedine_iniziali):
    for giocatore in lista_giocatori:
        truppe_rimanenti  = pedine_iniziali
        print(f"\nGiocatore {giocatore}, hai {truppe_rimanenti} truppe da posizionare.")
        territori_giocatore, _ = visualizza_stati_numero(giocatore)
        for territorio in territori_giocatore:
            aggiorna_truppe_stato(giocatore, territorio, 1)
            truppe_rimanenti -= 1
        
        print(f"✅ Ogni tuo territorio ha ora 1 truppa. Te ne rimangono {truppe_rimanenti} da posizionare.")

        while truppe_rimanenti > 0:
            i = random.randint(0, len(territori_giocatore) - 1)
            stato_scelto=territori_giocatore[i]
            territori_giocatore.remove(stato_scelto)
            numero_truppe=random.randint(1,truppe_rimanenti)
            aggiorna_truppe_stato(giocatore, stato_scelto, numero_truppe)
            truppe_rimanenti -= numero_truppe

        print(f"✅ Tutte le truppe sono state posizionate per il giocatore {giocatore}.")

# ------------------- ASSEGNA TRUPPE PER TERRITORI -------------------

def assegna_turno_truppe(giocatore, wb, continenti):

    territori_giocatore , count= visualizza_stati_numero(giocatore)

    pedine_da_posizionare = count // 3
    print(f"Il giocatore {giocatore} ha {count} territori, quindi spettano {pedine_da_posizionare} pedine.")

    if all(t in territori_giocatore for t in continenti["europa"]):
        pedine_da_posizionare += 5
        print(f"✅ {giocatore} ha conquistato tutti i territori in Europa e guadagna 5 pedine extra!")

    if all(t in territori_giocatore for t in continenti["asia"]):
        pedine_da_posizionare += 7
        print(f"✅ {giocatore} ha conquistato tutti i territori in Asia e guadagna 7 pedine extra!")

    if all(t in territori_giocatore for t in continenti["america_del_nord"]):
        pedine_da_posizionare += 5
        print(f"✅ {giocatore} ha conquistato tutti i territori in America del Nord e guadagna 5 pedine extra!")

    if all(t in territori_giocatore for t in continenti["america_del_sud"]):
        pedine_da_posizionare += 2
        print(f"✅ {giocatore} ha conquistato tutti i territori in America del Sud e guadagna 2 pedine extra!")

    if all(t in territori_giocatore for t in continenti["africa"]):
        pedine_da_posizionare += 3
        print(f"✅ {giocatore} ha conquistato tutti i territori in Africa e guadagna 3 pedine extra!")

    if all(t in territori_giocatore for t in continenti["oceania"]):
        pedine_da_posizionare += 2
        print(f"✅ {giocatore} ha conquistato tutti i territori in Oceania e guadagna 2 pedine extra!")
    
    inserisci_truppe(giocatore, pedine_da_posizionare, wb)


def inserisci_truppe(giocatore, pedine_da_posizionare, wb):
    truppe_posizionate = 0
    territori_giocatore , _ = visualizza_stati_numero(giocatore)
    while truppe_posizionate < pedine_da_posizionare:
        i = random.randint(0,len(territori_giocatore)-1)
        dove = territori_giocatore[i]
        quante = random.randint(1, pedine_da_posizionare - truppe_posizionate)
        territori_giocatore.remove(dove)
        # Trova la riga del territorio e scrivi le truppe in colonna D in questo modo non vengono sovrascrtte ma aggiunte
        aggiorna_truppe_stato(giocatore,dove,quante)

        truppe_posizionate += quante
        print(f"✅ Posizionate {quante} truppe in {dove} ({truppe_posizionate}/{pedine_da_posizionare})")

    wb.save(EXCEL_PATH)
    print("✅ Tutte le truppe sono state salvate nel file Excel.")

# ------------------- VISUALIZZA STATI (NO MAIN) -------------------

def visualizza_stati_numero(giocatore):
    print(f"visualizzo stati di {giocatore}")
    ws = wb[giocatore]
    n_territori = 0
    territori_giocatore = []
    for riga in range(9, 30):  # righe dove ci sono i territori
        nome_territorio = ws[f"A{riga}"].value
        if nome_territorio:  # <-- meglio controllare che non sia None
            territori_giocatore.append(nome_territorio)
            n_territori += 1
    return territori_giocatore, n_territori

# ------------------- VISUALIZZA ARMATE PER STATI (NO MAIN) -------------------

def trova_truppe_riga_stato(giocatore, stato):
    print(f"🔍 Cerco truppe per {giocatore} in {stato}")
    ws = wb[giocatore]
    for riga in range(9, 30):
        if ws[f"A{riga}"].value == stato:
            return ws[f"D{riga}"].value, riga  # restituisco anche la riga
    return 0, None
# ------------------- VISUALIZZA CARTE  -------------------

def trova_carte(giocatore):
    print(f"🔍 Cerco carte per {giocatore}")
    ws = wb[giocatore]
    lista_carte_giocatore = []
    for riga in range(9,30):
        carta= ws[f"C{riga}"].value
        if carta:
            lista_carte_giocatore.append(carta)
    return lista_carte_giocatore

# ------------------- VISUALIZZA GIOCATORI CON PAESE (NO MAIN) -------------------
def trova_giocatore(lista_giocatori, paese):
    print(f"🔍 Cerco giocatore in {paese}")
    for giocatore in lista_giocatori:
        ws = wb[giocatore]
        for riga in range(9, 30):  # <-- così controlli tutte le righe da 9 a 22
            if ws[f"A{riga}"].value == paese:
                return giocatore
    return None  # se non trovato


# ------------------- AGGIORNA NUMERO TRUPPE (NO MAIN) -------------------
def aggiorna_truppe_stato(giocatore, stato, n_truppe_aggiornate):
    print(f"🔄 Aggiorno truppe per {giocatore} in {stato}")
    ws = wb[giocatore]
    _, riga = trova_truppe_riga_stato(giocatore, stato)
    
    # Controlla se la riga è stata trovata
    if riga is None:
        print(f"❌ Errore: Stato '{stato}' non trovato per il giocatore '{giocatore}'.")
        return
        
    valore_casella_attuale = ws[f"D{riga}"].value
    if valore_casella_attuale is None:
        valore_casella_attuale = 0
    try:
        valore_casella_attuale = int(valore_casella_attuale)
    except (ValueError, TypeError):
        print(f"⚠️ Attenzione: Il valore nella cella D{riga} non è un numero valido. Impostato a 0.")
        valore_casella_attuale = 0

    ws[f"D{riga}"].value = valore_casella_attuale + n_truppe_aggiornate

    print(f"✅ {giocatore} ora ha in {stato}: {ws[f'D{riga}'].value}")
    wb.save(EXCEL_PATH)
    print("file aggiornato con successo")

# ------------------- PASSAGGIO STATO (NO MAIN) -------------------

# Codice corretto
def passaggio_stato(donatore, beneficiario, stato, n_truppe_attaccanti):
    ws_donatore = wb[donatore]
    ws_beneficiario = wb[beneficiario]

    _, riga_donatore = trova_truppe_riga_stato(donatore, stato)
    
    if riga_donatore is None:
        print(f"❌ {stato} non trovato tra i territori di {donatore}")
        return

    n_truppe_spostate = random.randint(1, n_truppe_attaccanti - 1)

    # Read the data from the donor's sheet
    nome = ws_donatore[f"A{riga_donatore}"].value
    continente = ws_donatore[f"B{riga_donatore}"].value
    simbolo = ws_donatore[f"C{riga_donatore}"].value

    print("Remove the territory and troops from the donor")
    for col in ["A", "B", "C", "D"]:
        ws_donatore[f"{col}{riga_donatore}"].value = None

    # Add the territory and troops to the beneficiary
    ws_beneficiario.append([nome, continente, simbolo, n_truppe_spostate])

    # Now, update the attacking player's original territory
    aggiorna_truppe_stato(beneficiario, stato, n_truppe_spostate)
    aggiorna_truppe_stato(donatore, stato, -n_truppe_spostate) # Remove from the original territory

    wb.save(EXCEL_PATH)
    print(f"✅ {stato} trasferito da {donatore} a {beneficiario} con successo, con {n_truppe_spostate} truppe.")

# ------------------- ATTACCO -------------------

def attacco(lista_giocatori, giocatore):
    territori_giocatore, _ = visualizza_stati_numero(giocatore)
    
    stato_partenza = random.choice(territori_giocatore)
    
    stato_attacco = None
    difensore = None

    territori_nemici = []
    for altro_giocatore in lista_giocatori:
        if altro_giocatore != giocatore:
            territori_nemici.extend(visualizza_stati_numero(altro_giocatore)[0])
            
    if not territori_nemici:
        print("❌ Non ci sono territori da attaccare. Salto il turno.")
        return
        
    stato_attacco = random.choice(territori_nemici)
    print(f"stato d'attacco {stato_attacco}")
    difensore = trova_giocatore(lista_giocatori, stato_attacco)
    print(f"difensore {difensore}")
    numero_truppe_stato_attaccante, _ = trova_truppe_riga_stato(giocatore, stato_partenza)
    numero_truppe_stato_difensore, _ = trova_truppe_riga_stato(difensore, stato_attacco)

    if numero_truppe_stato_attaccante <= 1:
        print(f"⚠️ {giocatore} non ha abbastanza truppe nello stato {stato_partenza}.")
        return

    # scelta dadi attaccante
    while True:
        n_dadi_attaccante = random.randint(1,3)
        print(f"🎲 {giocatore} lancia {n_dadi_attaccante} dadi")
        if n_dadi_attaccante >= numero_truppe_stato_attaccante:
            print(f"⚠️ Hai solo {numero_truppe_stato_attaccante} truppe in {stato_partenza}, "
                  f"quindi puoi tirare al massimo {numero_truppe_stato_attaccante - 1} dadi.")
            continue
        break

    # scelta dadi difensore
    while True:
        n_dadi_difensore = random.randint(1,2)
        if n_dadi_difensore > numero_truppe_stato_difensore:
            print(f"⚠️ {difensore} ha solo {numero_truppe_stato_difensore} truppe "
                  f"in {stato_attacco}, quindi può tirare al massimo {numero_truppe_stato_difensore} dadi.")
            continue
        break

    print(f"✅ {giocatore} attacca con {n_dadi_attaccante} dadi")
    print(f"✅ {difensore} difende con {n_dadi_difensore} dadi")

    
    # Attaccante
    dadi_attaccante = []
    for i in range(n_dadi_attaccante):
        dado = random.randint(1,6)
        print(f"🎲 {giocatore} ha lanciato un {dado}")
        dadi_attaccante.append(dado)

    # Difensore
    dadi_difensore = []
    for i in range(n_dadi_difensore):
        dado = random.randint(1,6)
        print(f"🎲 {difensore} ha lanciato un {dado}")
        dadi_difensore.append(dado)

    # Ordina i dadi dal più alto al più basso
    dadi_attaccante.sort(reverse=True)
    dadi_difensore.sort(reverse=True)

    # Confronta i primi "n" dadi
    n_confronti = min(len(dadi_attaccante), len(dadi_difensore))

    perdite_attaccante = 0
    perdite_difensore = 0

    for i in range(n_confronti):
        if dadi_attaccante[i] > dadi_difensore[i]:
            perdite_difensore += 1
        else:
            perdite_attaccante += 1

    print(f"{giocatore} perde {perdite_attaccante} truppe")
    print(f"{difensore} perde {perdite_difensore} truppe")

    aggiorna_truppe_stato(giocatore, stato_partenza, -perdite_attaccante)
    aggiorna_truppe_stato(difensore, stato_attacco, -perdite_difensore)

    numero_truppe_stato_difensore , _ = trova_truppe_riga_stato(difensore, stato_attacco)

    if numero_truppe_stato_difensore == 0:
        print(f"{difensore} non ha più truppe nello stato {stato_attacco} il paese passa a {giocatore}")
        passaggio_stato(difensore, giocatore, stato_attacco, numero_truppe_stato_attaccante)

# ------------------- ELIMINAZIONE GIOCATORE -------------------

def eliminazione_giocatore(lista_giocatori, turno):
    lista_sopravvissuti = []
    for giocatore in lista_giocatori:
        _ , n_territori = visualizza_stati_numero(giocatore)
        if n_territori == 0:
            print(f"Il giocatore {giocatore} è stato SCONFITTO al turno {turno} in quanto non possiede più territori")
        else:
            lista_sopravvissuti.append(giocatore)
    return lista_sopravvissuti

# ------------------- VINCITORE -------------------

def vincitore(giocatore, continenti):
    ws = wb[giocatore]
    obiettivo_assegnato = ws["B5"].value
    territori_giocatore, numero_stati = visualizza_stati_numero(giocatore)

    if obiettivo_assegnato == "Conquistare l'Europa, il Sud America e un terzo continente a scelta.":
        
        if all(t in territori_giocatore for t in continenti["europa"] + continenti["america_del_sud"]):
            print(f"{giocatore} ha conquistato Europa e Sud America!")

            altri_continenti = ["africa", "asia", "oceania", "america_del_nord"]
            for cont in altri_continenti:
                if all(t in territori_giocatore for t in continenti[cont]):
                    print(f"{giocatore} ha conquistato anche {cont}, obiettivo completato!")
                    return giocatore, True
    
    elif obiettivo_assegnato == "Conquistare l'Europa, l'Oceania e un terzo continente a scelta.":

        if all(t in territori_giocatore for t in continenti["europa"] + continenti["oceania"]):
            print(f"{giocatore} ha conquistato Europa e oceania!")

            altri_continenti = ["africa", "asia", "america_del_sud", "america_del_nord"]
            for cont in altri_continenti:
                if all(t in territori_giocatore for t in continenti[cont]):
                    print(f"{giocatore} ha conquistato anche {cont}, obiettivo completato!")
                    return giocatore, True    

    elif obiettivo_assegnato == "Conquistare l'Asia e il Sud America.":
        if all(t in territori_giocatore for t in continenti["asia"] + continenti["america_del_sud"]):
            print(f"{giocatore} ha conquistato asia e Sud America!")
            return giocatore, True  

    elif obiettivo_assegnato == "Conquistare l'Asia e l'Africa.":
        if all(t in territori_giocatore for t in continenti["asia"] + continenti["africa"]):
            print(f"{giocatore} ha conquistato asia e africa!")
            return giocatore, True  

    elif obiettivo_assegnato == "Conquistare il Nord America e l'Africa.":
        if all(t in territori_giocatore for t in continenti["america_del_nord"] + continenti["africa"]):
            print(f"{giocatore} ha conquistato nord america e africa!")
            return giocatore, True  

    elif obiettivo_assegnato == "Conquistare il Nord America e l'Oceania.":
        if all(t in territori_giocatore for t in continenti["america_del_nord"] + continenti["oceania"]):
            print(f"{giocatore} ha conquistato nord america e oceania!")
            return giocatore, True    

    elif obiettivo_assegnato == "Conquistare il Nord America e l'Europa.":
        if all(t in territori_giocatore for t in continenti["america_del_nord"] + continenti["europa"]):
            print(f"{giocatore} ha conquistato nord america e europa!")
            return giocatore, True

    elif obiettivo_assegnato == "Conquistare l'Africa e l'Europa.":
        if all(t in territori_giocatore for t in continenti["africa"] + continenti["europa"]):
            print(f"{giocatore} ha conquistato africa e europa!")
            return giocatore, True

    elif numero_stati == 24:
            print(f"{giocatore} ha conquistato 24 territori!")
            return giocatore, True

    elif numero_stati == 18:
            print(f"{giocatore} ha conquistato 18 territori!")
            return giocatore, True  

    return giocatore ,False

# ------------------- MAIN -------------------

if __name__ == "__main__":
    # ------------------- INIZIALIZZAZIONE -------------------
    lista_giocatori, n_giocatori = aggiungi_giocatori()
    lista_colori = assegna_colore(lista_giocatori)
    pedine_iniziali = pedine_distart(lista_giocatori)

    file_territori = carica_territori()
    file_obiettivi = carica_obiettivi()
    file_continenti = carica_continenti()
    
    wb, ws = verifica_file(lista_giocatori, lista_colori, pedine_iniziali)
    assegna_obiettivi(lista_giocatori, wb, file_obiettivi)
    numero_carte = assegna_territori(file_territori, lista_giocatori, wb)
    inserire_truppe_iniziali(lista_giocatori, pedine_iniziali)

    # ------------------- CICLO DI GIOCO -------------------
    turno = 0
    while len(lista_giocatori) > 1:
        print(f"\n=== Turno {turno + 1} ===")
        giocatore = lista_giocatori[turno % len(lista_giocatori)]
        print(f"\n--- È il turno di {giocatore} ---")
        assegna_turno_truppe(giocatore, wb, file_continenti)

        lista_carte_giocatore = trova_carte(giocatore)

        conteggio = Counter(lista_carte_giocatore)

        if  conteggio["fante"] >= 3 or conteggio["cannone"] >= 3 or conteggio["cavallo"] >= 3 or (conteggio["fante"] >=1 and conteggio["cannone"] >=1 and conteggio["cavallo"] >=1):
            scelta_carte = "s"
            if scelta_carte == "s":
                quale_tris = random.randint(1, 4)
                if quale_tris == 1:
                    if conteggio["fante"] >= 3:
                        print("benissimo hai 3 fanti nella lista")
                        inserisci_truppe(giocatore, 6, wb)
                        for _ in range(3):
                            lista_carte_giocatore.remove("fante")
                    else:
                        print("non ha abbastanaza fanti nella lista")
                elif quale_tris == 2:
                    if conteggio["cannone"] >= 3:
                        print("benissimo hai 3 cannoni nella lista")
                        inserisci_truppe(giocatore, 4, wb)
                        for _ in range(3):
                            lista_carte_giocatore.remove("cannone")
                    else:
                        print("non ha abbastanaza cannoni nella lista")
                elif quale_tris == 3:
                    if conteggio["cavallo"] >= 3:
                        print("benissimo hai 3 cavalli nella lista")
                        inserisci_truppe(giocatore, 10, wb)
                        for _ in range(3):
                            lista_carte_giocatore.remove("cavallo")
                    else:
                        print("non ha abbastanaza cavalli nella lista")
                elif quale_tris == 4:
                    if conteggio["fante"] >= 1 and conteggio["cannone"] >= 1 and conteggio["cavallo"] >= 1:
                        print("benissimo hai un fante, un cannone e un cavallo nella lista")
                        inserisci_truppe(giocatore, 10, wb)
                        lista_carte_giocatore.remove("fante")
                        lista_carte_giocatore.remove("cannone")
                        lista_carte_giocatore.remove("cavallo")
                    else:
                        print("non ha abbastanaza carte per il tris misto")

        while True:
            # scelta azione
            scelta_azione = random.randint(1, 3)

            # ------------------- PASSA TURNO -------------------
            if scelta_azione == 1:
                print(f"{giocatore} ha terminato il turno.\n")
                break

            # ------------------- ATTACCO -------------------
            elif scelta_azione == 2:
                print("\n--- Modalità attacco ---")
                while True:
                    attacco(lista_giocatori, giocatore)
                    continua_attacco = random.randint(1,2)
                    if continua_attacco != 1:
                        break
                print("--- Fine attacco ---")

            # ------------------- SPOSTAMENTO TRUPPE -------------------
            elif scelta_azione == 3:
                print("\n--- Modalità spostamento truppe ---")
                while True:
                    stati_disponibili, _ = visualizza_stati_numero(giocatore)
                    while True:
                        i_1 = random.randint(0, len(stati_disponibili) - 1)
                        stato_donatore = stati_disponibili[i_1]
                        truppe_disponibili, _ = trova_truppe_riga_stato(giocatore, stato_donatore)

                        # Aggiungi questa condizione
                        if truppe_disponibili > 1:
                            break  # Esci dal ciclo una volta trovato un territorio con più di 1 truppa
                    
                    # Ora puoi selezionare il territorio di destinazione in modo sicuro
                    while True:
                        i_2 = random.randint(0, len(stati_disponibili) - 1)
                        stato_beneficiario = stati_disponibili[i_2]
                        if stato_beneficiario != stato_donatore:
                            break
                            
                    n_truppe_da_posizionare = random.randint(1, truppe_disponibili - 1)

                    aggiorna_truppe_stato(giocatore, stato_donatore, -n_truppe_da_posizionare)
                    aggiorna_truppe_stato(giocatore, stato_beneficiario, n_truppe_da_posizionare)

                    continua_spostamento = random.randint(1,2)
                    if continua_spostamento != 1:
                        break

                print("--- Fine spostamento truppe ---")

        # ------------------- ELIMINAZIONE GIOCATORI -------------------
        lista_giocatori = eliminazione_giocatore(lista_giocatori, turno)
        if len(lista_giocatori) == 1:
            print(f"🎉 Il vincitore è {lista_giocatori[0]}!")
            break
        giocatore_vittorioso , vittoria = vincitore(giocatore,file_continenti)
        if vittoria == True:
            print(f"ABBIAMO UN VINCITORE : {giocatore_vittorioso}")
            break

        turno += 1
