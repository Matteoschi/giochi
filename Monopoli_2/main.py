import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import random
import json
import random

# === PERCORSI E COSTANTI ===
PATRIMONIO_INIZIALE = 10000
FOLDER = "Monopoli_2"
os.makedirs(FOLDER, exist_ok=True)
NOME_FILE_EXCEL = "monopoli_spese.xlsx"
EXCEL_PATH = os.path.join(FOLDER, NOME_FILE_EXCEL)
IMPREVISTI_PATH = r"C:\Users\alessandrini\Documents\coding\games\Monopoli_2\imprevisti.json"
BOARD_PATH = r"C:\Users\alessandrini\Documents\coding\games\Monopoli_2\monopoli_board.json"


def verifica_file(lista_giocatori):
    try:
        # Se i file board e imprevisti non esistono, non possiamo proseguire
        if not os.path.exists(IMPREVISTI_PATH) or not os.path.exists(BOARD_PATH):
            print("❌ File 'imprevisti.json' o 'monopoli_board.json' mancanti nella cartella.")
            return False, None, None

        if os.path.exists(EXCEL_PATH):
            print("📁 I file esistono!")
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            return True, wb, ws
        else:
            print("❌ File mancanti. Creo nuovo file Excel...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Giocatori"
            ws.append(["Nome Giocatore", "Saldo Iniziale"])

            # Aggiungi giocatori al foglio principale
            for giocatore in lista_giocatori:
                ws.append([giocatore, PATRIMONIO_INIZIALE])

            # Crea un foglio per ogni giocatore e inizializza con il saldo
            for giocatore in lista_giocatori:
                foglio = wb.create_sheet(title=giocatore)
                foglio.append(["Turno", "Casella", "ID Casella", "Tipo", "Colore", "Importo", "Saldo", "Descrizione"])
                # Aggiungi la riga iniziale con il patrimonio
                foglio.append([1, "Via", 0, "speciale", "NA", 0, PATRIMONIO_INIZIALE, "Inizio partita"])

            wb.save(EXCEL_PATH)
            print(f"✅ File '{NOME_FILE_EXCEL}' creato con i fogli dei giocatori.")
            return False, wb, ws
    
    except IndentationError:
        print(f"❌ Errore: Il file '{NOME_FILE_EXCEL}' non è un file Excel valido.")
        return False, None, None
    except Exception as e:
        print(f"❌ Errore nel controllo dei file: {e}")
        return False, None, None


def carica_imprevisti():
    try:
        with open(IMPREVISTI_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    except json.JSONDecodeError:
        print("❌ Errore nel file board.")
        return []
    
def carica_board():
    try:
        with open(BOARD_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    except json.JSONDecodeError:
        print("❌ Errore nel file board.")
        return []

def aggiungi_giocatori():
    lista_giocatori = []
    try:
        num_players = int(input("Quanti giocatori? "))
    except ValueError:
        print("❌ Inserisci un numero valido.")
        return aggiungi_giocatori()

    for i in range(num_players):
        while True:
            nome = input(f"Inserisci nome del giocatore {i+1}: ").strip().lower()
            if not nome:
                print("❌ Nome non valido.")
            elif nome in lista_giocatori:
                print("⚠️ Giocatore già inserito.")
            else:
                lista_giocatori.append(nome)
                print("✅ Giocatore aggiunto.")
                break
    print(f"🎮 Giocatori registrati: {lista_giocatori}")
    return lista_giocatori  


def lancia_dadi(giocatore, posizione_corrente, board, wb, imprevisti):
    try:
        dado = int(input(f"{giocatore}, lancia i dadi (2-12): "))
        if not 2 <= dado <= 12:
            print("❌ Inserisci un numero tra 2 e 12.")
            return lancia_dadi(giocatore, posizione_corrente, board, wb, imprevisti)
    except ValueError:
        print("❌ Inserisci un numero valido.")
        return lancia_dadi(giocatore, posizione_corrente, board, wb, imprevisti)

    nuova_posizione = (posizione_corrente + dado) % len(board)
    casella = next((p for p in board if p["posizione"] == nuova_posizione), None)

    if not casella:
        print("❌ Casella non trovata.")
        return posizione_corrente
    
    nome_casella = casella["nome"]
    id_casella = casella["posizione"]
    tipo_casella = casella.get("tipo")
    colore_casella = casella.get("colore")
    affitto = casella.get("affitto", 0)
    proprietario_casella = casella.get("acquistato")
    prezzo_casella = casella.get("prezzo", 0)
    stato_ipoteca = casella.get("stato_di_ipoteca")


    if giocatore not in wb.sheetnames:
        print(f"⚠️ Foglio per {giocatore} non trovato.")
        return nuova_posizione

    foglio = wb[giocatore]
    turno = foglio.max_row
    saldo = foglio.cell(row=turno, column=7).value if turno > 1 else PATRIMONIO_INIZIALE
    saldo = saldo if saldo is not None else PATRIMONIO_INIZIALE
    importo = 0


     #passaggio per il via
    if nuova_posizione < posizione_corrente:
        saldo += 200
        print(f"🎉 Passi dal Via e ricevi 200€! Saldo attuale: {saldo}€.")
        foglio.append([turno, "VIA !!", "0", "BONUS", "", 200, saldo, "passa dal via"])

    print(f"🎲 {giocatore} ha tirato {dado} e si sposta da {board[posizione_corrente]['nome']} a {casella['nome']}| tipo: {tipo_casella} | colore: {colore_casella} | proprietario: {proprietario_casella}")

    if tipo_casella == "proprietà":
        # COMPRARE LA CASE PER PRIMA VOLTA
        if not proprietario_casella and not stato_ipoteca:
            scelta = input(f"🏠 Vuoi acquistare {nome_casella} per {prezzo_casella}€? (s/n): ").strip().lower()
            if scelta == "s":
                if saldo >= prezzo_casella:
                    saldo -= prezzo_casella
                    casella["acquistato"] = giocatore
                    foglio.append([turno, nome_casella, id_casella, tipo_casella, colore_casella, -prezzo_casella, saldo, "acquisto proprietà"])  # segnare il colore della casella solo quando si compra per la prima volta
                    print(f"✅ Hai acquistato {nome_casella} per {prezzo_casella}€.")
                    with open(BOARD_PATH, "w", encoding="utf-8") as file:
                        json.dump(board, file, indent=4, ensure_ascii=False)
                    wb.save(EXCEL_PATH)
                else:
                    print("💸 Fondi insufficienti.")
            else:
                print("⏭ Hai deciso di non acquistare.")


        #PAGARE L'AFFITTO

        elif giocatore != proprietario_casella and not stato_ipoteca:
            print(f"💼 {giocatore} è atterrato su una proprietà di {proprietario_casella}.")
            if proprietario_casella in wb.sheetnames:
                foglio_prop = wb[proprietario_casella]
                turno_prop = foglio_prop.max_row
                saldo_prop = foglio_prop.cell(row=turno_prop, column=7).value if turno_prop > 1 else PATRIMONIO_INIZIALE
                saldo_prop = saldo_prop if saldo_prop is not None else PATRIMONIO_INIZIALE

                colori = {c: 0 for c in ["marrone", "celeste", "rosa", "arancione", "rosso", "giallo", "verde", "blu"]}
                for row in foglio_prop.iter_rows(min_row=2, min_col=5, max_col=5):
                    colore = str(row[0].value).strip().lower() if row[0].value else ""
                    if colore in colori:
                        colori[colore] += 1

                gruppo_completo = (
                    (colore_casella == "marrone" and colori["marrone"] == 2) or
                    (colore_casella == "celeste" and colori["celeste"] == 3) or
                    (colore_casella == "rosa" and colori["rosa"] == 3) or
                    (colore_casella == "arancione" and colori["arancione"] == 3) or
                    (colore_casella == "rosso" and colori["rosso"] == 3) or
                    (colore_casella == "giallo" and colori["giallo"] == 3) or
                    (colore_casella == "verde" and colori["verde"] == 3) or
                    (colore_casella == "blu" and colori["blu"] == 2)
                )

                numero_case = casella.get("numero_case")
                possiede_hotel = casella.get("possiede_hotel")

                if gruppo_completo and numero_case is None:
                    affitto_effettivo = affitto * 2
                elif not gruppo_completo :
                    affitto_effettivo = affitto
                elif numero_case is not None and not possiede_hotel and gruppo_completo:
                    if numero_case == "1":
                        affitto_effettivo = casella.get(f"affitto_{numero_case}_casa", 0)
                    else:
                        affitto_effettivo = casella.get(f"affitto_{numero_case}_case", 0)
                elif possiede_hotel:
                    affitto_effettivo = casella.get("affitto_albergo", 0)

                saldo -= affitto_effettivo
                saldo_prop += affitto_effettivo

                print(f"💸 {giocatore} paga {affitto_effettivo}€ a {proprietario_casella}")

                foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -affitto_effettivo, saldo, f"pagamento affitto a {proprietario_casella}"])
                foglio_prop.append([turno_prop, nome_casella, id_casella, tipo_casella, "NA", affitto_effettivo, saldo_prop, f"incasso affitto da {giocatore}"])
                wb.save(EXCEL_PATH)
            else:
                print(f"⚠️ Foglio del proprietario {proprietario_casella} non trovato.")



        # STESSO PROPRIETARIO - CASE - HOTEL
        elif giocatore == proprietario_casella and not stato_ipoteca:
            colori = {c: 0 for c in ["marrone", "celeste", "rosa", "arancione", "rosso", "giallo", "verde", "blu"]}
            for row in foglio.iter_rows(min_row=2, min_col=5, max_col=5):
                colore = str(row[0].value).strip().lower() if row[0].value else ""
                if colore in colori:
                    colori[colore] += 1

            gruppo_completo = (
                (colore_casella == "marrone" and colori["marrone"] == 2) or
                (colore_casella == "celeste" and colori["celeste"] == 3) or
                (colore_casella == "rosa" and colori["rosa"] == 3) or
                (colore_casella == "arancione" and colori["arancione"] == 3) or
                (colore_casella == "rosso" and colori["rosso"] == 3) or
                (colore_casella == "giallo" and colori["giallo"] == 3) or
                (colore_casella == "verde" and colori["verde"] == 3) or
                (colore_casella == "blu" and colori["blu"] == 2)
            )

            if gruppo_completo:
                try:
                    scelta = input("🏗️ Vuoi costruire? (1-4 per case, h per hotel): ").strip().lower()
                    if scelta in ["1", "2", "3", "4"]:
                        scelta = int(scelta)
                        if scelta == "1":
                            chiave_affitto = f"affitto_{scelta}_casa"
                        else:
                            chiave_affitto = f"affitto_{scelta}_case"

                        costo_casa = casella.get(chiave_affitto, 0)
                        if saldo >= costo_casa:
                            saldo -= costo_casa
                            casella["numero_case"] = scelta
                            foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -costo_casa, saldo, f"acquisto {scelta} casa/e"])
                            with open(BOARD_PATH, "w", encoding="utf-8") as file:
                                json.dump(board, file, indent=4, ensure_ascii=False)
                            wb.save(EXCEL_PATH)
                            print(f"✅ Hai costruito {scelta} casa/e per {costo_casa}€.")
                        else:
                            print("❌ Saldo insufficiente per costruire.")
                    elif scelta == "h" and casella.get("numero_case") == 4:
                        costo_hotel = casella.get("affitto_albergo", 0)
                        if saldo >= costo_hotel:
                            saldo -= costo_hotel
                            casella["possiede_hotel"] = True
                            foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -costo_hotel, saldo, "acquisto hotel"])
                            with open(BOARD_PATH, "w", encoding="utf-8") as file:
                                json.dump(board, file, indent=4, ensure_ascii=False)
                            wb.save(EXCEL_PATH)
                            print(f"✅ Hai costruito un hotel per {costo_hotel}€.")
                        else:
                            print("❌ Saldo insufficiente per costruire hotel.")
                    else:
                        print("❌ Scelta non valida o non hai abbastanza case per costruire un hotel.")
                except ValueError:
                    print("❌ Inserisci un numero valido.")
    
    if tipo_casella == "tassa":
        importo = casella.get("premio", 0)
        saldo -= importo
        foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -importo, saldo, "pagamento tassa"])
        print(f"{giocatore} ha pagato {importo} euro per essere capitato {nome_casella}")
        wb.save(EXCEL_PATH)

    if tipo_casella == "stazione":
        # COMPRARE LA STAZIONE
        if not proprietario_casella and not stato_ipoteca:
            scelta = input(f"🏠 Vuoi acquistare {nome_casella} per {prezzo_casella}€? (s/n): ").strip().lower()
            if scelta == "s":
                if saldo >= prezzo_casella:
                    saldo -= prezzo_casella
                    casella["acquistato"] = giocatore
                    foglio.append([turno, nome_casella, id_casella, tipo_casella, "Stazione", -prezzo_casella, saldo, "acquisto proprietà"]) 
                    print(f"✅ Hai acquistato {nome_casella} per {prezzo_casella}€.")
                    with open(BOARD_PATH, "w", encoding="utf-8") as file:
                        json.dump(board, file, indent=4, ensure_ascii=False)
                    wb.save(EXCEL_PATH)
                    print(f"il {giocatore} ha comprato per {prezzo_casella} {nome_casella}")
                else:
                    print("💸 Fondi insufficienti.")
            else:
                print("⏭ Hai deciso di non acquistare.")

        # PAGARE AFFITTO
        if proprietario_casella:
            if giocatore != proprietario_casella:
                print(f"🔁 Devi pagare {affitto}€ a {proprietario_casella}")
                
                if proprietario_casella in wb.sheetnames:
                    foglio_prop = wb[proprietario_casella]
                    turno_prop = foglio_prop.max_row
                    saldo_prop = foglio_prop.cell(row=turno_prop, column=7).value if turno_prop > 1 else PATRIMONIO_INIZIALE
                    saldo_prop = saldo_prop if saldo_prop is not None else PATRIMONIO_INIZIALE

                    # Conta quante stazioni possiede il proprietario
                    numero_stazioni = 0
                    for row in foglio_prop.iter_rows(min_row=2, min_col=5, max_col=5):
                        cell_value = row[0].value
                        if isinstance(cell_value, str) and "stazione" in cell_value.lower():
                            numero_stazioni += 1

                    affitto_totale = affitto * numero_stazioni
                    saldo -= affitto_totale
                    saldo_prop += affitto_totale

                    print(f"💸 {giocatore} paga {affitto_totale}€ a {proprietario_casella} (possiede {numero_stazioni} stazioni)")

                    foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -affitto_totale, saldo, f"pagamento affitto a {proprietario_casella}"])
                    foglio_prop.append([turno_prop, nome_casella, id_casella, tipo_casella, "NA", affitto_totale, saldo_prop, f"incasso affitto da {giocatore}"])
                    wb.save(EXCEL_PATH)

                else:
                    print(f"⚠️ Il foglio per il proprietario {proprietario_casella} non esiste!")

    if tipo_casella == "imprevisto":
        imprevisto = random.choice(imprevisti)
        saldo += imprevisto["premio"]
        foglio.append([turno, nome_casella, imprevisto["id"], tipo_casella, "NA", imprevisto["premio"], saldo, "imprevisto"])
        print(f"sei capitato in un imprevisto : {imprevisto['premio']} euro sul tuo conto")
        wb.save(EXCEL_PATH)

    if tipo_casella == "società":
        if not proprietario_casella and not stato_ipoteca:
            if saldo >= prezzo_casella:
                scelta_società=input(f"{giocatore} , vuoi comprare la società :{nome_casella} per {prezzo_casella} euro ? (s/n) ")
                if scelta_società == "s":
                        saldo -= prezzo_casella
                        casella["acquistato"] = giocatore
                        foglio.append([turno, nome_casella, id_casella, tipo_casella, "società", -prezzo_casella, saldo, "acquisto società"])
                        with open(BOARD_PATH, "w", encoding="utf-8") as file:
                            json.dump(board, file, indent=4, ensure_ascii=False)
                        wb.save(EXCEL_PATH)
                        print(f"complimenti {giocatore} hai comprato la società per {prezzo_casella} euro")
            else:
                print("non hai abbastanza soldi")


        if giocatore != proprietario_casella and proprietario_casella:
            print(f"💼 {giocatore} è atterrato su una proprietà di {proprietario_casella}.")
            if proprietario_casella in wb.sheetnames:
                foglio_prop = wb[proprietario_casella]
                turno_prop = foglio_prop.max_row
                saldo_prop = foglio_prop.cell(row=turno_prop, column=7).value if turno_prop > 1 else PATRIMONIO_INIZIALE
                saldo_prop = saldo_prop if saldo_prop is not None else PATRIMONIO_INIZIALE
            
            numero_società = 0
            for row in foglio_prop.iter_rows(min_row=1, min_col=5, max_col=5):
                cell_value = row[0].value
                if isinstance(cell_value, str) and cell_value.strip().lower() == "società":
                    numero_società += 1
            
            if numero_società > 1:
                importo = dado * 10
                print(f"{giocatore} paga {importo} euro perché il proprietario possiede 2 società (dado * 10)")
            else:
                importo = dado * 4
                print(f"{giocatore} paga {importo} euro perché il proprietario possiede 1 società (dado * 4)")
            
            saldo -= importo
            saldo_prop += importo
            foglio.append([turno, nome_casella, id_casella, tipo_casella, "NA", -importo, saldo, f"pagamento società a {proprietario_casella}"])
            foglio_prop.append([turno, nome_casella, id_casella, tipo_casella, "Na", importo, saldo_prop, f"ottenuto pagamento società da {giocatore} "])
            wb.save(EXCEL_PATH)


    return nuova_posizione


def main():
    if not os.path.exists(EXCEL_PATH):
        lista_giocatori = aggiungi_giocatori()
    else:
        lista_giocatori = None  

    file_esistente, wb, ws = verifica_file(lista_giocatori)

    if file_esistente:
        lista_giocatori = [
            row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]
        ]
        print(f"👥 Giocatori caricati dal file: {lista_giocatori}")

    imprevisti = carica_imprevisti()
    board = carica_board()
    posizioni = {g: 0 for g in lista_giocatori}

    giocatore_corrente = 0
    while True:
        giocatore = lista_giocatori[giocatore_corrente]
        posizioni[giocatore] = lancia_dadi(giocatore, posizioni[giocatore], board, wb, imprevisti)
        giocatore_corrente = (giocatore_corrente + 1) % len(lista_giocatori)

if __name__ == "__main__":
    main()

# TODO ipotecare  , scambio carte , famo tipo append colore 