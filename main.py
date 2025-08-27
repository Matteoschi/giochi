import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import random
import json
import random

territori = r"C:\Users\alessandrini\Documents\coding\games\territori.json"
obiettivi = r"Risiko\obiettivi.json"
FOLDER = "Risiko"
os.makedirs(FOLDER, exist_ok=True)
NOME_FILE_EXCEL = "RisiKo.xlsx"
EXCEL_PATH = os.path.join(FOLDER, NOME_FILE_EXCEL)


def verifica_file(lista_giocatori, lista_colori):
    try:
        # ✅ Controllo file JSON necessari
        if not os.path.exists(territori) or not os.path.exists(obiettivi):
            print("❌ File 'territori.json' o 'obiettivi.json' mancanti nella cartella.")
            return None, None

        # ✅ Controllo file Excel
        if os.path.exists(EXCEL_PATH):
            print("Il file Excel esiste, lo apro...")
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            return wb, ws
        else:
            print(" File Excel mancante, creo nuovo...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Giocatori"

            # Scrivo giocatore e colore corrispondente
            for giocatore, colore in zip(lista_giocatori, lista_colori):
                ws.append([giocatore, colore])
            
            wb.save(EXCEL_PATH)
            print(f"✅ File '{EXCEL_PATH}' creato con i fogli dei giocatori.")
            return wb, ws

    except Exception as e:
        print(f"❌ Errore nel controllo dei file: {e}")
        return None, None

def aggiungi_giocatori():
    lista_giocatori = []
    while True:
        try:
            n_giocatori = int(input("Inserisci numero giocatori (3-6): "))
            if 3 <= n_giocatori <= 6:
                break
            else:
                print("⚠️ I giocatori devono essere almeno 3 e massimo 6.")
        except ValueError:
            print("❌ Inserisci un numero valido.")

    for i in range(n_giocatori):
        while True:
            nome = input(f"Inserisci nome giocatore {i+1}: ").strip().capitalize()
            if not nome:
                print("⚠️ Inserire un nome valido.")
            elif nome in lista_giocatori:
                print("⚠️ Nome già inserito, scegline un altro.")
            else:
                lista_giocatori.append(nome)
                print(f"✅ Giocatore '{nome}' aggiunto.")
                break

    print(f"\n🎮 Giocatori registrati: {lista_giocatori}")
    return lista_giocatori , n_giocatori

def assegna_colore(lista_giocatori):
    lista_colori=[]
    for giocatore in lista_giocatori:
        while True:
            colore = input(f"inserisci colore per {giocatore}").strip().lower()
            if colore not in ["giallo", "rosso", "verde", "blu", "viola", "nero"]:
                print("colore non valido")
            elif colore in lista_colori:
                print("colore già esistente")
            else:
                lista_colori.append(colore)
                break

    return lista_colori

if __name__ == "__main__":
    lista_giocatori, n_giocatori = aggiungi_giocatori()
    lista_colori = assegna_colore(lista_giocatori)
    wb, ws = verifica_file(lista_giocatori, lista_colori)